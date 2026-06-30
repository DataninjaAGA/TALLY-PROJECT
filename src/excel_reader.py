from __future__ import annotations

from collections import defaultdict
from io import BytesIO
from typing import BinaryIO, Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .models import TallyDocument, TallyTask
from .utils import (
    clean_text,
    excel_serial_to_date,
    excel_time_to_hhmm,
    is_aircraft_register,
    normalize_register,
    one_line,
    should_exclude_task,
    wo_date_matches_tally_date,
    clean_work_order_for_tally,
)

# Canonical columns in the Daily Check sheet after the header row is found.
COL_AIRCRAFT = 1
COL_FLIGHT = 2
COL_ARR = 3
COL_DEPT = 4
COL_WO = 5
COL_TASK_CARD = 6
COL_DESCRIPTION = 7
COL_STATUS = 8
COL_DONE = 9
COL_PENDING = 10
COL_COMMENTS = 11
COL_MAN_HOURS = 12


class DailyCheckError(Exception):
    pass


def load_daily_check(uploaded_file: BinaryIO | bytes | str):
    """Load workbook with formula results where available."""
    if isinstance(uploaded_file, bytes):
        return load_workbook(BytesIO(uploaded_file), data_only=True)
    return load_workbook(uploaded_file, data_only=True)


def get_candidate_sheet_names(workbook) -> list[str]:
    """Return sheets that look like Daily Check operational sheets."""
    candidates: list[str] = []
    for ws in workbook.worksheets:
        header_row = find_header_row(ws)
        if header_row:
            candidates.append(ws.title)
    return candidates


def find_header_row(ws: Worksheet) -> int | None:
    """Find the row containing A/C, WO, TASK CARD, DESCRIPTION and STATUS."""
    max_scan = min(ws.max_row, 20)
    for row_idx in range(1, max_scan + 1):
        values = [one_line(ws.cell(row_idx, col).value).upper() for col in range(1, min(ws.max_column, 15) + 1)]
        joined = "|".join(values)
        if "A/C" in values and "WO" in values and "TASK CARD" in joined and "DESCRIPTION" in values:
            return row_idx
    return None


def extract_station(ws: Worksheet) -> str:
    # Prefer the value immediately after STA in the first rows.
    for row_idx in range(1, min(ws.max_row, 5) + 1):
        for col_idx in range(1, min(ws.max_column, 8) + 1):
            if one_line(ws.cell(row_idx, col_idx).value).upper() == "STA":
                return one_line(ws.cell(row_idx, col_idx + 1).value).upper()
    return ws.title.upper()


def extract_date(ws: Worksheet):
    # In the provided sample, the real date is an Excel serial near the DATE label.
    for row_idx in range(1, min(ws.max_row, 5) + 1):
        for col_idx in range(1, min(ws.max_column, 12) + 1):
            if one_line(ws.cell(row_idx, col_idx).value).upper() == "DATE":
                for offset in range(1, 8):
                    value = ws.cell(row_idx, col_idx + offset).value
                    parsed = excel_serial_to_date(value)
                    if parsed:
                        return parsed
    # fallback: any serial date in first row
    for col_idx in range(1, min(ws.max_column, 12) + 1):
        parsed = excel_serial_to_date(ws.cell(1, col_idx).value)
        if parsed:
            return parsed
    return None


def _build_man_hours_lookup(workbook) -> dict[str, str]:
    """Build TC -> Man Hours lookup from the HM sheet if present."""
    if "HM" not in workbook.sheetnames:
        return {}
    ws = workbook["HM"]
    lookup: dict[str, str] = {}
    for row_idx in range(2, ws.max_row + 1):
        task_card = one_line(ws.cell(row_idx, 1).value).upper()
        man_hours = one_line(ws.cell(row_idx, 3).value)
        if task_card and man_hours:
            lookup[task_card] = man_hours
    return lookup


def parse_daily_check(
    workbook,
    sheet_name: str | None = None,
    include_cancelled: bool = False,
    include_man_hours: bool = False,
) -> list[TallyDocument]:
    """Parse a Daily Check workbook into one TallyDocument per aircraft.

    Rules based on the sample package:
    - A new aircraft starts when column A matches a register, e.g. N506VL or XA-VRI.
    - Rows inherit WO from the last non-empty WO inside the aircraft group.
    - Cancelled/closed rows are omitted by default.
    - Aircraft groups without task card/description are ignored.
    """
    if sheet_name is None:
        candidates = get_candidate_sheet_names(workbook)
        if not candidates:
            raise DailyCheckError("No encontré una hoja con estructura de Daily Check.")
        sheet_name = candidates[0]
    if sheet_name not in workbook.sheetnames:
        raise DailyCheckError(f"La hoja '{sheet_name}' no existe en el archivo.")

    ws = workbook[sheet_name]
    header_row = find_header_row(ws)
    if not header_row:
        raise DailyCheckError(f"La hoja '{sheet_name}' no tiene encabezados reconocibles de Daily Check.")

    station = extract_station(ws)
    tally_date = extract_date(ws)
    mh_lookup = _build_man_hours_lookup(workbook)

    documents: list[TallyDocument] = []
    current: TallyDocument | None = None
    current_wo = ""
    current_wo_is_eligible = True

    for row_idx in range(header_row + 1, ws.max_row + 1):
        a_value = ws.cell(row_idx, COL_AIRCRAFT).value
        if is_aircraft_register(a_value):
            if current and current.tasks:
                documents.append(current)
            current = TallyDocument(
                register=normalize_register(a_value),
                station=station,
                tally_date=tally_date,
                flight=one_line(ws.cell(row_idx, COL_FLIGHT).value),
                arrival=excel_time_to_hhmm(ws.cell(row_idx, COL_ARR).value),
                departure=excel_time_to_hhmm(ws.cell(row_idx, COL_DEPT).value),
            )
            current_wo = ""

        if current is None:
            continue

        # Stop if we reached summary blocks rather than task data.
        if one_line(a_value).upper() in {"TAREAS", "RON AC", "AVIONES", "TOTAL"}:
            break

        wo_cell = one_line(ws.cell(row_idx, COL_WO).value)
        task_card = clean_text(ws.cell(row_idx, COL_TASK_CARD).value)
        description = clean_text(ws.cell(row_idx, COL_DESCRIPTION).value)
        status = clean_text(ws.cell(row_idx, COL_STATUS).value)
        done_value = ws.cell(row_idx, COL_DONE).value
        comments = clean_text(ws.cell(row_idx, COL_COMMENTS).value)
        tc_mh = one_line(ws.cell(row_idx, COL_MAN_HOURS).value)

        if wo_cell and not wo_cell.upper().startswith("TRANSIT CHECK"):
            current_wo = wo_cell
            current_wo_is_eligible = wo_date_matches_tally_date(wo_cell, tally_date)
        elif wo_cell and not current_wo and task_card:
            current_wo = wo_cell
            current_wo_is_eligible = wo_date_matches_tally_date(wo_cell, tally_date)

        # Rows such as "TRANSIT CHECK / RON" without task card and description do not create PDFs.
        if not task_card and not description:
            continue

        # If a WO explicitly belongs to a different calendar day, skip its tasks.
        # This excludes old STORAGE/PARKING work orders such as JUN18 when the tally date is JUN19.
        if not current_wo_is_eligible:
            continue

        if should_exclude_task(done_value, include_cancelled=include_cancelled):
            continue

        # Avoid importing helper/calculation rows without a real task.
        if not task_card:
            task_card = wo_cell
        if not description:
            description = task_card

        if include_man_hours:
            man_hours = tc_mh if tc_mh and tc_mh != "#N/A" else mh_lookup.get(one_line(task_card).upper(), "")
        else:
            man_hours = ""

        # The reference PDFs put STATUS from Daily Check into the REMARK column.
        remark = status
        if comments and include_cancelled:
            remark = f"{remark}\n{comments}" if remark else comments

        current.tasks.append(
            TallyTask(
                task_card=task_card,
                description=description,
                man_hours=man_hours,
                work_order=clean_work_order_for_tally(current_wo or wo_cell),
                status="",
                logbook_pg="",
                remark=remark,
                source_row=row_idx,
            )
        )

    if current and current.tasks:
        documents.append(current)

    # Remove duplicates only if Excel contains repeated aircraft sections; preserve order.
    return documents


def summarize_documents(documents: Iterable[TallyDocument]) -> list[dict[str, object]]:
    return [
        {
            "Register": doc.register,
            "Tasks": len(doc.tasks),
            "Station": doc.station,
            "Date": doc.tally_date.isoformat() if doc.tally_date else "",
            "First WO": doc.tasks[0].work_order if doc.tasks else "",
        }
        for doc in documents
    ]
